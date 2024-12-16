import os
import csv
from PIL import Image
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# Euclidean distance function for RGB colors
def color_distance(color1, color2):
    return math.sqrt((color2[0] - color1[0]) ** 2 +
                     (color2[1] - color1[1]) ** 2 +
                     (color2[2] - color1[2]) ** 2)


# Function to check if the color is too close to any existing color
def is_color_too_close(new_color, existing_colors, threshold=20):
    for color in existing_colors:
        if color_distance(new_color, color) < threshold:
            return True
    return False


def get_unique_colors_in_images(folder_path, threshold=20):
    unique_colors = []  # List of unique colors
    count = 0

    # Iterate over all files in the folder
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            image_path = os.path.join(folder_path, filename)

            # Open the image
            with Image.open(image_path) as img:
                img = img.convert('RGB')  # Convert image to RGB if it isn't already
                pixels = list(img.getdata())  # Get pixel data

                # Add each pixel to the unique color list if it is not too close to an existing color
                for pixel in pixels:
                    if not is_color_too_close(pixel, unique_colors, threshold):
                        unique_colors.append(pixel)
        count += 1
        print(f"Processed Image #{count}")

    # Return the unique colors
    return unique_colors


def save_colors_to_excel(unique_colors, output_file):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Unique Colors"

    # Add headers
    ws['A1'] = 'Red'
    ws['B1'] = 'Green'
    ws['C1'] = 'Blue'

    # Set cell color for each unique color row
    for i, color in enumerate(unique_colors, start=2):  # Start at row 2, row 1 is for headers
        red, green, blue = color
        # Set RGB background color using the color in each row
        hex_color = f'{red:02X}{green:02X}{blue:02X}'
        cell_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        # Write the RGB values to the cells
        ws[f'A{i}'] = red
        ws[f'B{i}'] = green
        ws[f'C{i}'] = blue

        # Set the background color of the row
        ws[f'A{i}'].fill = cell_fill
        ws[f'B{i}'].fill = cell_fill
        ws[f'C{i}'].fill = cell_fill

    # Save the workbook
    wb.save(output_file)


# Example usage
folder_path = "/Users/oliviaforte/DH 285 Final Project/Code/TEST"
output_excel_file = 'filtered_unique_colors.xlsx'

# Get the unique colors with the proximity filter
unique_colors = get_unique_colors_in_images(folder_path, threshold=20)

# Save the unique colors to the Excel file
save_colors_to_excel(unique_colors, output_excel_file)

print(f'Total unique colors saved to Excel: {len(unique_colors)}')
print("Done!")
